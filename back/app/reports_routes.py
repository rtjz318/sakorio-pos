"""
Sales / Revenue Reports API

Uses existing order and order_item data. Only paid/completed orders;
excludes removed and cancelled items. For restaurant owner revenue analysis.
"""

from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO, StringIO
from typing import Annotated
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from . import models
from .db import get_session
from .permissions import Permission, require_permission
from .rate_limits import admin_user_limit
from .report_export_i18n import report_export_labels
from .security import get_current_user
from .work_session_serialization import serialize_work_session

router = APIRouter()

# Order states that count as collected revenue. Kitchen/service "completed"
# without a payment timestamp is not cash-up revenue.
REVENUE_STATUSES = {models.OrderStatus.paid}
# Item statuses we exclude from revenue
EXCLUDED_ITEM_STATUSES = {models.OrderItemStatus.cancelled}

_PAYMENT_METHOD_ORDER = {
    "hitpay": 0,
    "terminal": 1,
    "cash": 2,
    "other": 3,
}


def _as_utc(value: datetime) -> datetime:
    """Normalize legacy naive timestamps and current aware timestamps to UTC."""
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _revenue_date(order: models.Order) -> datetime | None:
    """Date used for attributing revenue (paid_at if set, else created_at)."""
    return order.paid_at or order.created_at


def _order_counts_as_revenue(order: models.Order) -> bool:
    """Cash-up revenue must be collected, not merely kitchen/service completed."""
    return order.deleted_at is None and (order.status == models.OrderStatus.paid or order.paid_at is not None)


def _tenant_report_timezone(session: Session, tenant_id: int):
    tenant = session.get(models.Tenant, tenant_id)
    timezone_name = (tenant.timezone or "UTC").strip() if tenant else "UTC"
    try:
        return ZoneInfo(timezone_name)
    except (ZoneInfoNotFoundError, ValueError):
        return timezone.utc


def _local_report_date(value: datetime | date | None, report_timezone) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return _as_utc(value).astimezone(report_timezone).date()
    return value


def _in_range(
    value: datetime | date | None,
    from_date: date,
    to_date: date,
    report_timezone,
) -> bool:
    local_date = _local_report_date(value, report_timezone)
    if local_date is None:
        return False
    return from_date <= local_date <= to_date


def _waiter_name_for_order_tips(session: Session, order: models.Order) -> str:
    """Waiter for tip attribution: stored tip_attributed_user_id, else table assignment."""
    tid = getattr(order, "tip_attributed_user_id", None)
    if tid:
        u = session.get(models.User, tid)
        if u and u.tenant_id == order.tenant_id:
            return u.full_name or u.email or str(tid)
    table = session.get(models.Table, order.table_id) if order.table_id is not None else None
    if table:
        waiter_id = table.assigned_waiter_id
        if waiter_id is None and table.floor_id:
            floor = session.get(models.Floor, table.floor_id)
            if floor:
                waiter_id = floor.default_waiter_id
        if waiter_id:
            u = session.get(models.User, waiter_id)
            return (u.full_name or u.email) if u else str(waiter_id)
    return "Unassigned"


def _normalize_payment_method(raw: str | None) -> str:
    value = (raw or "").strip().lower().replace("-", "_").replace(" ", "_")
    if value in {"card_terminal", "terminal", "card"}:
        return "terminal"
    if value in {"cash"}:
        return "cash"
    if value in {"hitpay"}:
        return "hitpay"
    return "other"


def _get_revenue_items(
    session: Session,
    tenant_id: int,
    from_date: date,
    to_date: date,
    report_timezone,
):
    """Load orders and items that count toward revenue in the date range."""
    orders = session.exec(
        select(models.Order)
        .where(models.Order.tenant_id == tenant_id)
        .where(models.Order.deleted_at.is_(None))
        .order_by(models.Order.created_at.asc())
    ).all()

    result = []
    for order in orders:
        if not _order_counts_as_revenue(order):
            continue
        rev_date = _revenue_date(order)
        if not _in_range(rev_date, from_date, to_date, report_timezone):
            continue
        items = session.exec(
            select(models.OrderItem)
            .where(models.OrderItem.order_id == order.id)
            .where(models.OrderItem.removed_by_customer == False)
            .where(models.OrderItem.status != models.OrderItemStatus.cancelled)
        ).all()
        table = session.get(models.Table, order.table_id) if order.table_id is not None else None
        waiter_id = None
        waiter_name = None
        if table:
            waiter_id = table.assigned_waiter_id
            if waiter_id is None and table.floor_id:
                floor = session.get(models.Floor, table.floor_id)
                if floor:
                    waiter_id = floor.default_waiter_id
            if waiter_id:
                u = session.get(models.User, waiter_id)
                waiter_name = (u.full_name or u.email) if u else str(waiter_id)
        table_name = table.name if table else "Unknown"
        for item in items:
            product = session.get(models.Product, item.product_id)
            category = (product.category or "Uncategorized") if product else "Uncategorized"
            subcategory = (product.subcategory or "") if product else ""
            unit_cost = getattr(item, "cost_cents", None) or 0
            revenue_cents = item.quantity * item.price_cents
            cost_cents = item.quantity * unit_cost
            result.append({
                "order_id": order.id,
                "date": rev_date,
                "table_id": order.table_id,
                "table_name": table_name,
                "waiter_id": waiter_id,
                "waiter_name": waiter_name or "Unassigned",
                "product_id": item.product_id,
                "product_name": item.product_name,
                "category": category,
                "subcategory": subcategory,
                "quantity": item.quantity,
                "price_cents": item.price_cents,
                "cost_cents": cost_cents,
                "revenue_cents": revenue_cents,
                "profit_cents": revenue_cents - cost_cents,
            })
    return result


def _build_report_payload(tenant_id: int, session: Session, from_date: date, to_date: date) -> dict:
    """Build full report dict for a tenant and date range."""
    if from_date > to_date:
        from_date, to_date = to_date, from_date
    report_timezone = _tenant_report_timezone(session, tenant_id)
    rows = _get_revenue_items(
        session,
        tenant_id,
        from_date,
        to_date,
        report_timezone,
    )

    tips_by_day: dict[str, int] = defaultdict(int)
    tips_by_waiter: dict[str, int] = defaultdict(int)
    total_tips_cents = 0
    paid_orders_summary: list[dict] = []
    orders_for_tips = session.exec(
        select(models.Order)
        .where(models.Order.tenant_id == tenant_id)
        .where(models.Order.deleted_at.is_(None))
    ).all()
    for order in orders_for_tips:
        if not _order_counts_as_revenue(order):
            continue
        rev_date = _revenue_date(order)
        if not _in_range(rev_date, from_date, to_date, report_timezone):
            continue
        order_items = session.exec(
            select(models.OrderItem)
            .where(models.OrderItem.order_id == order.id)
            .where(models.OrderItem.removed_by_customer == False)
            .where(models.OrderItem.status != models.OrderItemStatus.cancelled)
        ).all()
        order_revenue_cents = sum(item.quantity * item.price_cents for item in order_items)
        tip = int(order.tip_amount_cents or 0)
        payment_method = _normalize_payment_method(getattr(order, "payment_method", None))
        paid_orders_summary.append(
            {
                "payment_method": payment_method,
                "revenue_cents": order_revenue_cents,
                "tips_cents": tip,
                "collected_cents": order_revenue_cents + tip,
            }
        )
        if tip <= 0:
            continue
        total_tips_cents += tip
        day = _local_report_date(rev_date, report_timezone).isoformat()
        tips_by_day[day] += tip
        wn = _waiter_name_for_order_tips(session, order)
        tips_by_waiter[wn] += tip

    # Summary by day
    by_day_agg: dict[str, dict] = defaultdict(
        lambda: {"revenue_cents": 0, "cost_cents": 0, "profit_cents": 0, "order_count": set()}
    )
    for r in rows:
        day = _local_report_date(r["date"], report_timezone).isoformat()
        by_day_agg[day]["revenue_cents"] += r["revenue_cents"]
        by_day_agg[day]["cost_cents"] += r["cost_cents"]
        by_day_agg[day]["profit_cents"] += r["profit_cents"]
        by_day_agg[day]["order_count"].add(r["order_id"])
    summary_daily = [
        {
            "date": d,
            "revenue_cents": data["revenue_cents"],
            "cost_cents": data["cost_cents"],
            "profit_cents": data["profit_cents"],
            "order_count": len(data["order_count"]),
            "tips_cents": tips_by_day.get(d, 0),
        }
        for d, data in sorted(by_day_agg.items())
    ]
    total_revenue_cents = sum(r["revenue_cents"] for r in rows)
    total_cost_cents = sum(r["cost_cents"] for r in rows)
    total_profit_cents = total_revenue_cents - total_cost_cents
    total_orders = len(set(r["order_id"] for r in rows))

    # By product
    by_product: dict[tuple[int, str], dict] = defaultdict(
        lambda: {"quantity": 0, "revenue_cents": 0, "cost_cents": 0, "profit_cents": 0, "category": ""}
    )
    for r in rows:
        key = (r["product_id"], r["product_name"])
        by_product[key]["quantity"] += r["quantity"]
        by_product[key]["revenue_cents"] += r["revenue_cents"]
        by_product[key]["cost_cents"] += r["cost_cents"]
        by_product[key]["profit_cents"] += r["profit_cents"]
        if not by_product[key]["category"]:
            by_product[key]["category"] = r.get("category") or "Uncategorized"
    by_product_list = [
        {
            "product_id": k[0],
            "product_name": k[1],
            "category": v["category"],
            "quantity": v["quantity"],
            "revenue_cents": v["revenue_cents"],
            "cost_cents": v["cost_cents"],
            "profit_cents": v["profit_cents"],
        }
        for k, v in sorted(by_product.items(), key=lambda x: -x[1]["revenue_cents"])
    ]

    # By category
    by_category: dict[str, dict] = defaultdict(
        lambda: {"quantity": 0, "revenue_cents": 0, "cost_cents": 0, "profit_cents": 0}
    )
    for r in rows:
        c = r["category"] or "Uncategorized"
        by_category[c]["quantity"] += r["quantity"]
        by_category[c]["revenue_cents"] += r["revenue_cents"]
        by_category[c]["cost_cents"] += r["cost_cents"]
        by_category[c]["profit_cents"] += r["profit_cents"]
    by_category_list = [
        {
            "category": k,
            "quantity": v["quantity"],
            "revenue_cents": v["revenue_cents"],
            "cost_cents": v["cost_cents"],
            "profit_cents": v["profit_cents"],
        }
        for k, v in sorted(by_category.items(), key=lambda x: -x[1]["revenue_cents"])
    ]

    # By table
    by_table: dict[str, dict] = defaultdict(lambda: {"revenue_cents": 0, "cost_cents": 0, "profit_cents": 0, "order_count": set()})
    for r in rows:
        t = r["table_name"]
        by_table[t]["revenue_cents"] += r["revenue_cents"]
        by_table[t]["cost_cents"] += r["cost_cents"]
        by_table[t]["profit_cents"] += r["profit_cents"]
        by_table[t]["order_count"].add(r["order_id"])
    by_table_list = [
        {
            "table_name": k,
            "revenue_cents": v["revenue_cents"],
            "cost_cents": v["cost_cents"],
            "profit_cents": v["profit_cents"],
            "order_count": len(v["order_count"]),
        }
        for k, v in sorted(by_table.items(), key=lambda x: -x[1]["revenue_cents"])
    ]

    # By waiter
    by_waiter: dict[str, dict] = defaultdict(
        lambda: {"revenue_cents": 0, "cost_cents": 0, "profit_cents": 0, "order_count": set()}
    )
    for r in rows:
        w = r["waiter_name"]
        by_waiter[w]["revenue_cents"] += r["revenue_cents"]
        by_waiter[w]["cost_cents"] += r["cost_cents"]
        by_waiter[w]["profit_cents"] += r["profit_cents"]
        by_waiter[w]["order_count"].add(r["order_id"])
    by_waiter_list = [
        {
            "waiter_name": k,
            "revenue_cents": v["revenue_cents"],
            "cost_cents": v["cost_cents"],
            "profit_cents": v["profit_cents"],
            "order_count": len(v["order_count"]),
            "tips_cents": tips_by_waiter.get(k, 0),
        }
        for k, v in sorted(by_waiter.items(), key=lambda x: -x[1]["revenue_cents"])
    ]

    by_payment_method: dict[str, dict] = defaultdict(
        lambda: {"revenue_cents": 0, "tips_cents": 0, "collected_cents": 0, "order_count": 0}
    )
    total_collected_cents = 0
    for paid in paid_orders_summary:
        method = paid["payment_method"]
        by_payment_method[method]["revenue_cents"] += paid["revenue_cents"]
        by_payment_method[method]["tips_cents"] += paid["tips_cents"]
        by_payment_method[method]["collected_cents"] += paid["collected_cents"]
        by_payment_method[method]["order_count"] += 1
        total_collected_cents += paid["collected_cents"]
    by_payment_method_list = []
    for method, data in sorted(
        by_payment_method.items(),
        key=lambda item: (_PAYMENT_METHOD_ORDER.get(item[0], 99), -item[1]["collected_cents"]),
    ):
        order_count = data["order_count"]
        collected_cents = data["collected_cents"]
        by_payment_method_list.append(
            {
                "payment_method": method,
                "revenue_cents": data["revenue_cents"],
                "tips_cents": data["tips_cents"],
                "collected_cents": collected_cents,
                "order_count": order_count,
                "average_collected_per_order_cents": (
                    collected_cents // order_count if order_count else 0
                ),
                "share_of_collected_sales_pct": (
                    round((collected_cents / total_collected_cents) * 100, 2)
                    if total_collected_cents > 0
                    else 0
                ),
            }
        )

    # Reservations in date range (by reservation_date); source = public (token set) vs staff (no token); by status
    reservations = session.exec(
        select(models.Reservation)
        .where(models.Reservation.tenant_id == tenant_id)
        .where(models.Reservation.reservation_date >= from_date)
        .where(models.Reservation.reservation_date <= to_date)
    ).all()
    total_reservations = len(reservations)
    by_source: dict[str, int] = defaultdict(int)
    by_status: dict[str, int] = defaultdict(int)
    for r in reservations:
        source = "public" if r.token else "staff"
        by_source[source] += 1
        by_status[r.status.value] += 1

    # Overbooking: count slots (date, time) in range where reserved_guests > total_seats or reserved_parties > total_tables
    tables = session.exec(select(models.Table).where(models.Table.tenant_id == tenant_id)).all()
    total_seats = sum(t.seat_count for t in tables)
    total_tables = len(tables)
    active_reservations = [r for r in reservations if r.status in (models.ReservationStatus.booked, models.ReservationStatus.seated)]
    slot_aggregates: dict[tuple[date, time], tuple[int, int]] = defaultdict(lambda: (0, 0))  # (guests, parties)
    for r in active_reservations:
        key = (r.reservation_date, r.reservation_time)
        g, p = slot_aggregates[key]
        slot_aggregates[key] = (g + r.party_size, p + 1)
    overbooking_slots_count = sum(
        1 for (g, p) in slot_aggregates.values() if g > total_seats or p > total_tables
    )

    reservations_summary = {
        "total": total_reservations,
        "by_source": [
            {"source": k, "count": v} for k, v in sorted(by_source.items(), key=lambda x: -x[1])
        ],
        "by_status": [
            {"status": k, "count": v} for k, v in sorted(by_status.items(), key=lambda x: -x[1])
        ],
        "overbooking_slots_count": overbooking_slots_count,
    }

    # Queue entries in date range (by requested_at) for waitlist / walk-in analytics.
    queue_entries = session.exec(
        select(models.GuestQueueEntry)
        .where(models.GuestQueueEntry.tenant_id == tenant_id)
    ).all()
    queue_entries = [
        entry
        for entry in queue_entries
        if _in_range(entry.requested_at, from_date, to_date, report_timezone)
    ]
    queue_total = len(queue_entries)
    queue_by_source: dict[str, int] = defaultdict(int)
    queue_by_status: dict[str, int] = defaultdict(int)
    queue_daily: dict[str, dict] = defaultdict(lambda: {"count": 0, "seated_count": 0})
    quoted_wait_values: list[int] = []
    actual_wait_values: list[int] = []
    for entry in queue_entries:
        source_key = entry.source.value if hasattr(entry.source, "value") else str(entry.source or "unknown")
        status_key = entry.status.value if hasattr(entry.status, "value") else str(entry.status or "unknown")
        queue_by_source[source_key] += 1
        queue_by_status[status_key] += 1
        if entry.requested_at:
            day_key = _local_report_date(entry.requested_at, report_timezone).isoformat()
            queue_daily[day_key]["count"] += 1
            if status_key == models.GuestQueueStatus.seated.value:
                queue_daily[day_key]["seated_count"] += 1
        if entry.quoted_wait_minutes is not None:
            quoted_wait_values.append(int(entry.quoted_wait_minutes))
        if entry.requested_at and entry.seated_at:
            actual_wait_minutes = int(
                max(
                    0,
                    round(
                        (_as_utc(entry.seated_at) - _as_utc(entry.requested_at)).total_seconds()
                        / 60
                    ),
                )
            )
            actual_wait_values.append(actual_wait_minutes)

    queue_seated_count = queue_by_status.get(models.GuestQueueStatus.seated.value, 0)
    queue_converted_count = queue_by_status.get(models.GuestQueueStatus.converted_to_reservation.value, 0)
    queue_cancelled_count = queue_by_status.get(models.GuestQueueStatus.cancelled.value, 0)
    queue_no_show_count = queue_by_status.get(models.GuestQueueStatus.no_show.value, 0)
    queue_expired_count = queue_by_status.get(models.GuestQueueStatus.expired.value, 0)
    queue_waiting_count = queue_by_status.get(models.GuestQueueStatus.waiting.value, 0)
    queue_notified_count = queue_by_status.get(models.GuestQueueStatus.notified.value, 0)
    seated_or_converted_count = queue_seated_count + queue_converted_count

    queue_summary = {
        "total": queue_total,
        "waiting_count": queue_waiting_count,
        "notified_count": queue_notified_count,
        "seated_count": queue_seated_count,
        "converted_to_reservation_count": queue_converted_count,
        "cancelled_count": queue_cancelled_count,
        "no_show_count": queue_no_show_count,
        "expired_count": queue_expired_count,
        "average_quoted_wait_minutes": (
            round(sum(quoted_wait_values) / len(quoted_wait_values), 1)
            if quoted_wait_values
            else 0
        ),
        "average_actual_wait_minutes": (
            round(sum(actual_wait_values) / len(actual_wait_values), 1)
            if actual_wait_values
            else 0
        ),
        "seat_conversion_pct": (
            round((seated_or_converted_count / queue_total) * 100, 1)
            if queue_total > 0
            else 0
        ),
        "converted_to_reservation_pct": (
            round((queue_converted_count / queue_total) * 100, 1)
            if queue_total > 0
            else 0
        ),
        "by_source": [
            {"source": key, "count": value}
            for key, value in sorted(queue_by_source.items(), key=lambda x: -x[1])
        ],
        "by_status": [
            {"status": key, "count": value}
            for key, value in sorted(queue_by_status.items(), key=lambda x: -x[1])
        ],
        "daily": [
            {"date": day, "count": data["count"], "seated_count": data["seated_count"]}
            for day, data in sorted(queue_daily.items())
        ],
    }

    average_revenue_per_order_cents = (
        total_revenue_cents // total_orders if total_orders else 0
    )
    return {
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "summary": {
            "total_revenue_cents": total_revenue_cents,
            "total_cost_cents": total_cost_cents,
            "total_profit_cents": total_profit_cents,
            "total_tips_cents": total_tips_cents,
            "total_collected_cents": total_collected_cents,
            "total_orders": total_orders,
            "average_revenue_per_order_cents": average_revenue_per_order_cents,
            "daily": summary_daily,
        },
        "reservations": reservations_summary,
        "queue": queue_summary,
        "by_payment_method": by_payment_method_list,
        "by_product": by_product_list,
        "by_category": by_category_list,
        "by_table": by_table_list,
        "by_waiter": by_waiter_list,
    }


@router.get("/sales")
@admin_user_limit()
def get_sales_reports(
    request: Request,
    response: Response,
    current_user: Annotated[models.User, Depends(require_permission(Permission.REPORT_READ))],
    session: Session = Depends(get_session),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
) -> dict:
    """Combined sales report for the given date range. Uses paid/completed orders only."""
    return _build_report_payload(current_user.tenant_id, session, from_date, to_date)


def _csv_stream(rows: list[dict], keys: list[str], header_row: list[str]) -> bytes:
    import csv

    # csv.writer requires a text stream; BytesIO expects bytes and raises TypeError.
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(header_row)
    for r in rows:
        writer.writerow([r.get(k, "") for k in keys])
    # utf-8-sig so Excel recognizes UTF-8 when headers contain non-ASCII (localized exports).
    return buf.getvalue().encode("utf-8-sig")


@router.get("/export")
@admin_user_limit()
def export_report(
    request: Request,
    response: Response,
    current_user: Annotated[models.User, Depends(require_permission(Permission.REPORT_READ))],
    session: Session = Depends(get_session),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    format: str = Query("csv", description="csv or xlsx"),
    report: str = Query("summary", description="summary, products, category, table, waiter, payment"),
    lang: str | None = Query(None, description="UI language for headers (e.g. en, es, de)"),
) -> StreamingResponse:
    """Export report as CSV or Excel. Same date range as reports."""
    if from_date > to_date:
        from_date, to_date = to_date, from_date
    data = _build_report_payload(current_user.tenant_id, session, from_date, to_date)
    L = report_export_labels(lang)

    if format.lower() == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            from fastapi import HTTPException
            raise HTTPException(500, "Excel export requires openpyxl")
        wb = Workbook()
        # Summary sheet
        ws = wb.active
        ws.title = L["sheet_summary"][:31]
        ws.append(
            [
                L["date"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
                L["tips_cents"],
                L["collected_cents"],
                L["orders"],
            ]
        )
        for row in data["summary"]["daily"]:
            ws.append([
                row["date"],
                row["revenue_cents"],
                row.get("cost_cents", 0),
                row.get("profit_cents", 0),
                row.get("tips_cents", 0),
                row["revenue_cents"] + row.get("tips_cents", 0),
                row["order_count"],
            ])
        ws.append([])
        s = data["summary"]
        ws.append(
            [
                L["total"],
                s["total_revenue_cents"],
                s.get("total_cost_cents", 0),
                s.get("total_profit_cents", 0),
                s.get("total_tips_cents", 0),
                s.get("total_collected_cents", s["total_revenue_cents"] + s.get("total_tips_cents", 0)),
                s["total_orders"],
            ]
        )
        # Reservations
        res = data.get("reservations", {})
        ws_res = wb.create_sheet(L["sheet_reservations"][:31])
        ws_res.append([L["source"], L["count"]])
        for row in res.get("by_source", []):
            sk = row["source"]
            src_label = L.get(f"source_{sk}", sk)
            ws_res.append([src_label, row["count"]])
        ws_res.append([])
        ws_res.append([L["status"], L["count"]])
        for row in res.get("by_status", []):
            st = row["status"]
            st_label = L.get(f"res_status_{st}", st)
            ws_res.append([st_label, row["count"]])
        ws_res.append([])
        ws_res.append([L["total"], res.get("total", 0)])
        # Products
        ws2 = wb.create_sheet(L["sheet_by_product"][:31])
        ws2.append(
            [
                L["product"],
                L["category"],
                L["quantity"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
            ]
        )
        for p in data["by_product"]:
            ws2.append([
                p["product_name"],
                p.get("category", ""),
                p["quantity"],
                p["revenue_cents"],
                p.get("cost_cents", 0),
                p.get("profit_cents", 0),
            ])
        # Category
        ws3 = wb.create_sheet(L["sheet_by_category"][:31])
        ws3.append(
            [
                L["category"],
                L["quantity"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
            ]
        )
        for c in data["by_category"]:
            ws3.append([
                c["category"],
                c["quantity"],
                c["revenue_cents"],
                c.get("cost_cents", 0),
                c.get("profit_cents", 0),
            ])
        # Table
        ws4 = wb.create_sheet(L["sheet_by_table"][:31])
        ws4.append(
            [
                L["table"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
                L["orders"],
            ]
        )
        for t in data["by_table"]:
            ws4.append([
                t["table_name"],
                t["revenue_cents"],
                t.get("cost_cents", 0),
                t.get("profit_cents", 0),
                t["order_count"],
            ])
        # Waiter
        ws5 = wb.create_sheet(L["sheet_by_waiter"][:31])
        ws5.append(
            [
                L["waiter"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
                L["tips_cents"],
                L["orders"],
            ]
        )
        for w in data["by_waiter"]:
            ws5.append([
                w["waiter_name"],
                w["revenue_cents"],
                w.get("cost_cents", 0),
                w.get("profit_cents", 0),
                w.get("tips_cents", 0),
                w["order_count"],
            ])
        # Payment method
        ws6 = wb.create_sheet(L["sheet_by_payment_method"][:31])
        ws6.append(
            [
                L["payment_method"],
                L["orders"],
                L["revenue_cents"],
                L["tips_cents"],
                L["collected_cents"],
                L["average_ticket_cents"],
                L["share_of_collected_sales_pct"],
            ]
        )
        for pm in data["by_payment_method"]:
            method_key = pm["payment_method"]
            method_label = L.get(f"payment_method_{method_key}", method_key)
            ws6.append([
                method_label,
                pm["order_count"],
                pm["revenue_cents"],
                pm.get("tips_cents", 0),
                pm.get("collected_cents", 0),
                pm.get("average_collected_per_order_cents", 0),
                pm.get("share_of_collected_sales_pct", 0),
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=pos2-sales-{from_date}-{to_date}.xlsx"},
        )

    # CSV: single report type
    if report == "summary":
        rows = [
            {
                "date": r["date"],
                "revenue_cents": r["revenue_cents"],
                "cost_cents": r.get("cost_cents", 0),
                "profit_cents": r.get("profit_cents", 0),
                "tips_cents": r.get("tips_cents", 0),
                "order_count": r["order_count"],
            }
            for r in data["summary"]["daily"]
        ]
        keys = ["date", "revenue_cents", "cost_cents", "profit_cents", "tips_cents", "order_count"]
        header_row = [
            L["date"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
            L["tips_cents"],
            L["orders"],
        ]
    elif report == "products":
        rows = data["by_product"]
        keys = ["product_name", "category", "quantity", "revenue_cents", "cost_cents", "profit_cents"]
        header_row = [
            L["product"],
            L["category"],
            L["quantity"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
        ]
    elif report == "category":
        rows = data["by_category"]
        keys = ["category", "quantity", "revenue_cents", "cost_cents", "profit_cents"]
        header_row = [
            L["category"],
            L["quantity"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
        ]
    elif report == "table":
        rows = data["by_table"]
        keys = ["table_name", "revenue_cents", "cost_cents", "profit_cents", "order_count"]
        header_row = [
            L["table"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
            L["orders"],
        ]
    elif report == "waiter":
        rows = data["by_waiter"]
        keys = ["waiter_name", "revenue_cents", "cost_cents", "profit_cents", "tips_cents", "order_count"]
        header_row = [
            L["waiter"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
            L["tips_cents"],
            L["orders"],
        ]
    elif report == "payment":
        rows = data["by_payment_method"]
        keys = [
            "payment_method",
            "order_count",
            "revenue_cents",
            "tips_cents",
            "collected_cents",
            "average_collected_per_order_cents",
            "share_of_collected_sales_pct",
        ]
        header_row = [
            L["payment_method"],
            L["orders"],
            L["revenue_cents"],
            L["tips_cents"],
            L["collected_cents"],
            L["average_ticket_cents"],
            L["share_of_collected_sales_pct"],
        ]
        rows = [
            {
                **row,
                "payment_method": L.get(
                    f"payment_method_{row['payment_method']}", row["payment_method"]
                ),
            }
            for row in rows
        ]
    else:
        rows = data["summary"]["daily"]
        keys = ["date", "revenue_cents", "order_count"]
        header_row = [L["date"], L["revenue_cents"], L["orders"]]
    content = _csv_stream(rows, keys, header_row)
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=pos2-sales-{report}-{from_date}-{to_date}.csv"},
    )


@router.get("/work-sessions")
@admin_user_limit()
def report_work_sessions(
    request: Request,
    response: Response,
    current_user: Annotated[
        models.User,
        Depends(require_permission(Permission.REPORT_READ, Permission.PAYROLL_SUMMARY_READ)),
    ],
    session: Session = Depends(get_session),
    from_date: str = Query(..., description="Start date YYYY-MM-DD (UTC day)"),
    to_date: str = Query(..., description="End date YYYY-MM-DD (UTC day, inclusive)"),
    user_id: int | None = Query(None, description="Filter by staff user id (optional)"),
) -> list[dict]:
    """All tenant staff clock-in/out rows in range (by started_at, UTC). Owner and admin only."""
    if current_user.tenant_id is None:
        raise HTTPException(status_code=403, detail="Tenant required")
    try:
        fd = datetime.strptime(from_date, "%Y-%m-%d").date()
        td = datetime.strptime(to_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format; use YYYY-MM-DD")
    if fd > td:
        raise HTTPException(status_code=400, detail="from_date must be <= to_date")
    start_utc = datetime.combine(fd, time.min, tzinfo=timezone.utc)
    end_exclusive = datetime.combine(td + timedelta(days=1), time.min, tzinfo=timezone.utc)
    q = (
        select(models.WorkSession)
        .where(models.WorkSession.tenant_id == current_user.tenant_id)
        .where(models.WorkSession.started_at >= start_utc)
        .where(models.WorkSession.started_at < end_exclusive)
    )
    if user_id is not None:
        q = q.where(models.WorkSession.user_id == user_id)
    rows = session.exec(q.order_by(models.WorkSession.started_at.desc())).all()
    out: list[dict] = []
    for ws in rows:
        u = session.get(models.User, ws.user_id)
        name = (u.full_name or u.email or "") if u else ""
        out.append(serialize_work_session(ws, name, session=session, include_payroll=True))
    return out


class WorkSessionAdjustBody(BaseModel):
    """Owner/admin manual correction of recorded times."""

    note: str = ""
    started_at: datetime | None = None
    ended_at: datetime | None = None


@router.get("/work-sessions/live")
@admin_user_limit()
def report_work_sessions_live(
    request: Request,
    response: Response,
    current_user: Annotated[models.User, Depends(require_permission(Permission.REPORT_READ))],
    session: Session = Depends(get_session),
) -> list[dict]:
    """Open clock sessions for the tenant (who is on-site / on break)."""
    if current_user.tenant_id is None:
        raise HTTPException(status_code=403, detail="Tenant required")
    rows = session.exec(
        select(models.WorkSession)
        .where(models.WorkSession.tenant_id == current_user.tenant_id)
        .where(models.WorkSession.ended_at.is_(None))
        .order_by(models.WorkSession.started_at.asc())
    ).all()
    out: list[dict] = []
    for ws in rows:
        u = session.get(models.User, ws.user_id)
        name = (u.full_name or u.email or "") if u else ""
        d = serialize_work_session(ws, name, session=session)
        d["user_role"] = u.role.value if u else None
        out.append(d)
    return out


@router.post("/work-sessions/{work_session_id}/adjust")
@admin_user_limit()
def adjust_work_session_times(
    request: Request,
    response: Response,
    work_session_id: int,
    body: WorkSessionAdjustBody,
    current_user: Annotated[models.User, Depends(require_permission(Permission.REPORT_READ))],
    session: Session = Depends(get_session),
) -> dict:
    """Manual override of clock-in/out times (audit trail). Owner/admin via report permission."""
    if current_user.tenant_id is None:
        raise HTTPException(status_code=403, detail="Tenant required")
    ws = session.exec(
        select(models.WorkSession).where(
            models.WorkSession.id == work_session_id,
            models.WorkSession.tenant_id == current_user.tenant_id,
        )
    ).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Work session not found")
    if body.started_at is None and body.ended_at is None:
        raise HTTPException(status_code=400, detail="Provide started_at and/or ended_at")

    prev_s, prev_e = ws.started_at, ws.ended_at
    new_s = body.started_at if body.started_at is not None else ws.started_at
    new_e = body.ended_at if body.ended_at is not None else ws.ended_at
    if new_s is None:
        raise HTTPException(status_code=400, detail="started_at cannot be cleared")
    if new_e is not None and new_e < new_s:
        raise HTTPException(status_code=400, detail="ended_at must be on or after started_at")

    adj = models.WorkSessionAdjustment(
        tenant_id=ws.tenant_id,
        work_session_id=ws.id,
        actor_user_id=current_user.id,
        note=(body.note or "").strip(),
        previous_started_at=prev_s,
        previous_ended_at=prev_e,
        new_started_at=new_s,
        new_ended_at=new_e,
    )
    session.add(adj)
    ws.started_at = new_s
    ws.ended_at = new_e
    session.add(ws)
    session.commit()
    session.refresh(ws)
    u = session.get(models.User, ws.user_id)
    name = (u.full_name or u.email or "") if u else ""
    return serialize_work_session(ws, name, session=session)
